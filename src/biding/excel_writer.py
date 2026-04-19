# excel_writer.py — xlsx output writer for the Quoting Calculator.
#
# Responsibility: take a CalculationResult and append it to an xlsx workbook,
# prepending the new block at the top of today's sheet (creating the sheet
# and/or the file if needed), so the most recent run is always at the top.
#
# Only this module imports openpyxl. No argparse. No console output.

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException  # re-raised to caller

from biding.models import CalculationResult, QuoteStep


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_result(result: CalculationResult, output_path: Path) -> None:
    """Prepend one run block to today's sheet in the xlsx workbook at output_path.

    Behaviour:
      - If the file does not exist: create a new workbook, write the block.
      - If the file exists: open it, find/create today's sheet, prepend the block.
      - All other sheets are left untouched.
      - If output_path's parent directory does not exist it is created.

    Raises:
      InvalidFileException: if output_path points to a corrupt / non-xlsx file.
      OSError: on filesystem permission or I/O problems.
    """
    # Ensure the parent directory exists (creates intermediate dirs as needed)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Derive today's sheet name from the calculation timestamp
    sheet_name = result.calculation_time.strftime("%Y-%m-%d")

    # Load or create the workbook
    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
        # Remove the default empty sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Find or create today's sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Build the list of row-tuples for this run block
    block_rows = _render_block_rows(result)

    # Prepend the block (shifts existing rows down if sheet has content)
    _prepend_block(ws, block_rows)

    # Polish: set column widths to 20 for columns A–E if still at default
    _set_column_widths(ws)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Block rendering
# ---------------------------------------------------------------------------

def _render_block_rows(result: CalculationResult) -> list[list]:
    """Return a list of row-lists representing one run block.

    Layout (design doc 04-excel-format.md §3):
      Row 1:  ["Overview", ""]
      Row 2:  ["Calculation Time", "2026-04-19 10:30:00"]
      Row 3:  ["Start Price",      "100.00"]
      Row 4:  ["Target Price",     "45.00"]
      Row 5:  ["Max Reduction Pct","50.00 %"]
      Row 6:  ["Min Reduction",    "10.00"]
      Row 7:  ["Decimals",         "2"]
      Row 8:  ["Rounding",         "true"]
      Row 9:  ["", ""]   — blank separator
      Row 10: column headers (bold)
      Row 11+: one data row per QuoteStep
    """
    params = result.params
    rows: list[list] = []

    # Overview sub-section (8 data rows + 1 blank = 9 rows total)
    rows.append(["Overview", ""])
    rows.append([
        "Calculation Time",
        result.calculation_time.strftime("%Y-%m-%d %H:%M:%S"),
    ])
    rows.append(["Start Price", _fmt(params.start_price)])
    rows.append(["Target Price", _fmt(result.effective_target)])
    rows.append(["Max Reduction Pct", _fmt(params.max_pct) + " %"])
    rows.append(["Min Reduction", _fmt(params.min_reduction)])
    rows.append(["Decimals", str(params.decimals)])
    rows.append(["Rounding", "true" if params.rounding else "false"])
    rows.append(["", ""])  # blank separator (row index 8, 0-based)

    # Quoting Sequence sub-section
    rows.append(["Round", "Start Amount", "End Amount", "Reduction Amount", "Reduction Pct"])
    for step in result.steps:
        rows.append(_step_row(step))

    return rows


def _fmt(value: Decimal) -> str:
    """Format a Decimal as a fixed-point string, preventing scientific notation.

    Using format(value, "f") ensures Excel does not re-interpret the cell
    as a float and avoids precision loss on import.
    """
    return format(value, "f")


def _step_row(step: QuoteStep) -> list:
    """Convert one QuoteStep into a list of cell values (one per column)."""
    return [
        step.round_no,
        _fmt(step.start_amount),
        _fmt(step.end_amount),
        _fmt(step.reduction_amount),
        _fmt(step.reduction_pct) + " %",
    ]


# ---------------------------------------------------------------------------
# Sheet manipulation
# ---------------------------------------------------------------------------

def _prepend_block(ws, block_rows: list[list]) -> None:
    """Insert block_rows at the top of ws, shifting any existing content down.

    If the sheet is empty, write directly from row 1.
    Otherwise, insert (len(block_rows) + 2) rows at position 1 — the 2 extra
    rows become the blank separator between the new and previous blocks —
    then write block_rows into rows 1..len(block_rows).
    """
    # Detect whether the sheet already has content
    sheet_is_empty = ws.max_row == 1 and ws.cell(1, 1).value is None

    if not sheet_is_empty:
        # Shift existing content down to make room for new block + 2 blank separator rows
        shift = len(block_rows) + 2
        ws.insert_rows(idx=1, amount=shift)

    # Write block starting at row 1
    _write_rows(ws, block_rows, start_row=1)


def _write_rows(ws, rows: list[list], start_row: int) -> None:
    """Write row-lists into the worksheet beginning at start_row.

    Formatting applied:
      - Row offset 0  ("Overview" label): bold
      - Row offset 9  (column headers):   bold
    """
    bold = Font(bold=True)

    for row_offset, row_data in enumerate(rows):
        row_num = start_row + row_offset
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(row=row_num, column=col_offset + 1, value=value)

            # Bold the "Overview" section header
            if row_offset == 0 and col_offset == 0:
                cell.font = bold

            # Bold the Quoting Sequence column-header row (0-based index 9)
            if row_offset == 9:
                cell.font = bold


def _set_column_widths(ws) -> None:
    """Set columns A–E to width 20 if they are still at the openpyxl default."""
    for letter in ("A", "B", "C", "D", "E"):
        dim = ws.column_dimensions[letter]
        # openpyxl reports default (unset) width as None; set once if not customised
        if dim.width is None or dim.width <= 8.44:
            dim.width = 20
