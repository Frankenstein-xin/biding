# test_main.py — Integration tests for biding.main.main (CLI orchestration).

import pytest
import openpyxl

from biding.main import main


# ---------------------------------------------------------------------------
# Helper: build a minimal valid argv list
# ---------------------------------------------------------------------------

def _argv(
    start="100",
    target="45",
    max_pct="50",
    min_red="10",
    decimals="2",
    rounding="true",
    output: str | None = None,
) -> list[str]:
    if output is None:
        output = "/tmp/biding_test_output.xlsx"
    return [
        "--start-price", start,
        "--target-price", target,
        "--max-pct", max_pct,
        "--min-reduction", min_red,
        "--decimals", decimals,
        "--rounding", rounding,
        "--output", output,
    ]


# ---------------------------------------------------------------------------
# Happy path — exit 0, file produced
# ---------------------------------------------------------------------------

class TestHappyPath:
    def test_exit_code_zero(self, tmp_path):
        code = main(_argv(output=str(tmp_path / "quotes.xlsx")))
        assert code == 0

    def test_output_file_created(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        main(_argv(output=str(out)))
        assert out.exists()

    def test_output_has_sheet(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        main(_argv(output=str(out)))
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) == 1

    def test_stdout_contains_rounds(self, tmp_path, capsys):
        out = tmp_path / "quotes.xlsx"
        main(_argv(output=str(out)))
        captured = capsys.readouterr()
        assert "round" in captured.out.lower()

    def test_auto_target(self, tmp_path):
        out = str(tmp_path / "quotes.xlsx")
        code = main(_argv(
            start="1000", target="0", max_pct="10", min_red="5", output=out
        ))
        assert code == 0


# ---------------------------------------------------------------------------
# Missing required flag → exit 2 (argparse SystemExit)
# ---------------------------------------------------------------------------

class TestMissingFlag:
    def test_missing_start_price(self, tmp_path):
        argv = _argv(output=str(tmp_path / "q.xlsx"))
        # Drop --start-price and its value
        idx = argv.index("--start-price")
        argv = argv[:idx] + argv[idx + 2:]
        with pytest.raises(SystemExit) as exc_info:
            main(argv)
        assert exc_info.value.code == 2

    def test_missing_output(self, tmp_path):
        argv = _argv(output=str(tmp_path / "q.xlsx"))
        idx = argv.index("--output")
        argv = argv[:idx] + argv[idx + 2:]
        with pytest.raises(SystemExit) as exc_info:
            main(argv)
        assert exc_info.value.code == 2


# ---------------------------------------------------------------------------
# Validation error → exit 3
# ---------------------------------------------------------------------------

class TestValidationError:
    def test_negative_start_price(self, tmp_path):
        code = main(_argv(start="-10", output=str(tmp_path / "q.xlsx")))
        assert code == 3

    def test_target_equals_start(self, tmp_path):
        code = main(_argv(start="100", target="100", output=str(tmp_path / "q.xlsx")))
        assert code == 3

    def test_max_pct_zero(self, tmp_path):
        code = main(_argv(max_pct="0", output=str(tmp_path / "q.xlsx")))
        assert code == 3


# ---------------------------------------------------------------------------
# Infeasible inputs → exit 4
# ---------------------------------------------------------------------------

class TestInfeasible:
    def test_infeasible_gap(self, tmp_path):
        # start=10, target=5, gap=5 but min_reduction=10 → gap < min_reduction
        out = str(tmp_path / "q.xlsx")
        code = main(_argv(
            start="10", target="5", max_pct="10", min_red="10", output=out
        ))
        assert code == 4

    def test_no_file_written_on_infeasible(self, tmp_path):
        out = tmp_path / "q.xlsx"
        main(_argv(start="10", target="5", max_pct="10", min_red="10", output=str(out)))
        # File must NOT be created when the calculation is infeasible
        assert not out.exists()


# ---------------------------------------------------------------------------
# Two runs same day — both succeed, sheet count stays at 1
# ---------------------------------------------------------------------------

class TestTwoRunsSameDay:
    def test_second_run_exits_zero(self, tmp_path):
        out = str(tmp_path / "quotes.xlsx")
        assert main(_argv(output=out)) == 0
        assert main(_argv(output=out)) == 0

    def test_sheet_count_unchanged(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        main(_argv(output=str(out)))
        main(_argv(output=str(out)))
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) == 1
