# test_excel_writer.py — Unit tests for biding.excel_writer.write_result.

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from biding.excel_writer import write_result
from biding.models import CalculationParams, CalculationResult, QuoteStep


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_result(
    start: str = "100",
    target: str = "45",
    max_pct: str = "50",
    min_red: str = "10",
    decimals: int = 2,
    rounding: bool = True,
    output_path: Path = Path("/tmp/test.xlsx"),
    calc_time: datetime = datetime(2026, 4, 19, 10, 30, 0),
    steps: tuple | None = None,
    effective_target: str | None = None,
) -> CalculationResult:
    params = CalculationParams(
        start_price=Decimal(start),
        target_price=Decimal(target),
        max_pct=Decimal(max_pct),
        min_reduction=Decimal(min_red),
        decimals=decimals,
        rounding=rounding,
        output_path=output_path,
    )
    if steps is None:
        steps = (
            QuoteStep(
                round_no=1,
                start_amount=Decimal("100.00"),
                end_amount=Decimal("55.00"),
                reduction_amount=Decimal("45.00"),
                reduction_pct=Decimal("45.00"),
            ),
            QuoteStep(
                round_no=2,
                start_amount=Decimal("55.00"),
                end_amount=Decimal("45.00"),
                reduction_amount=Decimal("10.00"),
                reduction_pct=Decimal("18.18"),
            ),
        )
    if effective_target is None:
        effective_target = target
    return CalculationResult(
        params=params,
        calculation_time=calc_time,
        effective_target=Decimal(effective_target),
        steps=steps,
    )


# ---------------------------------------------------------------------------
# New file — basic structure
# ---------------------------------------------------------------------------

class TestNewFile:
    def test_creates_file(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out), out)
        assert out.exists()

    def test_sheet_name_is_date(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out, calc_time=datetime(2026, 4, 19, 10, 30, 0)), out)
        wb = openpyxl.load_workbook(out)
        assert "2026-04-19" in wb.sheetnames

    def test_overview_header_cell(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(1, 1).value == "Overview"

    def test_calculation_time_row(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out, calc_time=datetime(2026, 4, 19, 10, 30, 0)), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(2, 1).value == "Calculation Time"
        assert ws.cell(2, 2).value == "2026-04-19 10:30:00"

    def test_start_price_row(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(3, 1).value == "Start Price"
        assert ws.cell(3, 2).value == "100"

    def test_rounding_true(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out, rounding=True), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(8, 2).value == "true"

    def test_rounding_false(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out, rounding=False), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(8, 2).value == "false"

    def test_column_headers_row(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        # Row 10 is the column-header row (rows 1-8 overview, row 9 blank, row 10 headers)
        assert ws.cell(10, 1).value == "Round"
        assert ws.cell(10, 2).value == "Start Amount"
        assert ws.cell(10, 5).value == "Reduction Pct"

    def test_first_data_row(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(11, 1).value == 1          # round_no
        assert ws.cell(11, 2).value == "100.00"   # start_amount
        assert ws.cell(11, 3).value == "55.00"    # end_amount

    def test_only_one_sheet_created(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out), out)
        assert len(openpyxl.load_workbook(out).sheetnames) == 1


# ---------------------------------------------------------------------------
# Same-day second run — insert at top
# ---------------------------------------------------------------------------

class TestSameDaySecondRun:
    def test_second_block_starts_at_row_one(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        t = datetime(2026, 4, 19, 10, 30, 0)
        write_result(_make_result(output_path=out, calc_time=t), out)
        write_result(_make_result(output_path=out, calc_time=t), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(1, 1).value == "Overview"

    def test_previous_block_shifted_down(self, tmp_path):
        # Block has 2 steps → 10 header rows + 2 step rows = 12 rows total
        # shift = 12 + 2 = 14; old row 1 → new row 15
        out = tmp_path / "quotes.xlsx"
        t = datetime(2026, 4, 19, 10, 30, 0)
        write_result(_make_result(output_path=out, calc_time=t), out)
        write_result(_make_result(output_path=out, calc_time=t), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(15, 1).value == "Overview"

    def test_two_blank_rows_between_blocks(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        t = datetime(2026, 4, 19, 10, 30, 0)
        write_result(_make_result(output_path=out, calc_time=t), out)
        write_result(_make_result(output_path=out, calc_time=t), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        # Rows 13 and 14 must be blank (separator between new and old block)
        assert ws.cell(13, 1).value is None
        assert ws.cell(14, 1).value is None

    def test_sheet_count_unchanged_same_day(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        t = datetime(2026, 4, 19, 10, 30, 0)
        write_result(_make_result(output_path=out, calc_time=t), out)
        write_result(_make_result(output_path=out, calc_time=t), out)
        assert len(openpyxl.load_workbook(out).sheetnames) == 1


# ---------------------------------------------------------------------------
# Different day — new sheet
# ---------------------------------------------------------------------------

class TestDifferentDay:
    def test_new_sheet_for_new_day(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out, calc_time=datetime(2026, 4, 19, 10, 0, 0)), out)
        write_result(_make_result(output_path=out, calc_time=datetime(2026, 4, 20, 9, 0, 0)), out)
        wb = openpyxl.load_workbook(out)
        assert "2026-04-19" in wb.sheetnames
        assert "2026-04-20" in wb.sheetnames

    def test_previous_sheet_untouched(self, tmp_path):
        out = tmp_path / "quotes.xlsx"
        write_result(_make_result(output_path=out, calc_time=datetime(2026, 4, 19, 10, 0, 0)), out)
        write_result(_make_result(output_path=out, calc_time=datetime(2026, 4, 20, 9, 0, 0)), out)
        ws = openpyxl.load_workbook(out)["2026-04-19"]
        assert ws.cell(1, 1).value == "Overview"


# ---------------------------------------------------------------------------
# Nested output path — auto directory creation
# ---------------------------------------------------------------------------

class TestNestedOutputPath:
    def test_creates_parent_directories(self, tmp_path):
        out = tmp_path / "nested" / "dir" / "quotes.xlsx"
        write_result(_make_result(output_path=out), out)
        assert out.exists()


# ---------------------------------------------------------------------------
# Corrupt file — exception propagates
# ---------------------------------------------------------------------------

class TestCorruptFile:
    def test_corrupt_file_raises(self, tmp_path):
        out = tmp_path / "bad.xlsx"
        out.write_bytes(b"this is not a valid xlsx file")
        with pytest.raises(Exception):
            write_result(_make_result(output_path=out), out)
