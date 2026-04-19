"""End-to-end and unit tests for the biding quoting calculator."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from biding.calculator import auto_target, calculate
from biding.excel_writer import write_result
from biding.main import main
from biding.models import (
    CalculationParams,
    CalculationResult,
    InfeasibleError,
    QuoteStep,
)
from biding.rounding import quantize


def _params(**overrides) -> CalculationParams:
    base = dict(
        start_price=Decimal("100"),
        target_price=Decimal("45"),
        max_pct=Decimal("50"),
        min_reduction=Decimal("10"),
        decimals=2,
        rounding=True,
        output_path=Path("/tmp/unused.xlsx"),
    )
    base.update(overrides)
    return CalculationParams(**base)


# --- rounding ---------------------------------------------------------------


def test_quantize_round_half_up() -> None:
    assert quantize(Decimal("1.235"), 2, True) == Decimal("1.24")


def test_quantize_truncate() -> None:
    assert quantize(Decimal("1.239"), 2, False) == Decimal("1.23")


def test_quantize_preserves_scale() -> None:
    assert str(quantize(Decimal("2"), 2, True)) == "2.00"


# --- models -----------------------------------------------------------------


def test_params_valid() -> None:
    assert _params().start_price == Decimal("100")


@pytest.mark.parametrize(
    "overrides, message",
    [
        ({"start_price": Decimal("0")}, "start-price must be > 0"),
        ({"target_price": Decimal("-1")}, "target-price must be >= 0"),
        ({"max_pct": Decimal("0")}, "max-pct must be in (0, 100)"),
        ({"max_pct": Decimal("100")}, "max-pct must be in (0, 100)"),
        ({"min_reduction": Decimal("0")}, "min-reduction must be > 0"),
        ({"decimals": -1}, "decimals must be >= 0"),
        (
            {"start_price": Decimal("10"), "target_price": Decimal("20")},
            "target-price must be < start-price",
        ),
    ],
)
def test_params_validation(overrides, message) -> None:
    with pytest.raises(ValueError) as exc:
        _params(**overrides)
    assert message in str(exc.value)


# --- calculator -------------------------------------------------------------


def test_calculate_direct_hit() -> None:
    r = calculate(_params(target_price=Decimal("50")))
    assert len(r.steps) == 1
    assert r.steps[0].end_amount == Decimal("50.00")


def test_calculate_prd_section_7_trap() -> None:
    # Naive "always max reduction" dead-ends; algorithm takes 100→55→45.
    r = calculate(_params())
    prices = [r.steps[0].start_amount] + [s.end_amount for s in r.steps]
    assert prices == [Decimal("100.00"), Decimal("55.00"), Decimal("45.00")]


def test_calculate_auto_target() -> None:
    p = _params(
        start_price=Decimal("1000"),
        target_price=Decimal("0"),
        max_pct=Decimal("10"),
        min_reduction=Decimal("5"),
    )
    assert auto_target(p) == Decimal("50.00")
    r = calculate(p)
    assert r.effective_target == Decimal("50.00")
    for step in r.steps:
        assert step.reduction_amount >= p.min_reduction
        assert step.reduction_amount <= step.start_amount * p.max_pct / Decimal("100")
    assert r.steps[-1].end_amount == Decimal("50.00")


def test_calculate_infeasible() -> None:
    with pytest.raises(InfeasibleError):
        calculate(
            _params(
                start_price=Decimal("10"),
                target_price=Decimal("5"),
                max_pct=Decimal("10"),
                min_reduction=Decimal("10"),
            )
        )


def test_calculate_rounding_off_lands_on_target() -> None:
    r = calculate(_params(rounding=False))
    assert r.steps[-1].end_amount == Decimal("45.00")


# --- excel_writer -----------------------------------------------------------


def _result(params: CalculationParams, when: datetime) -> CalculationResult:
    return CalculationResult(
        params=params,
        calculation_time=when,
        effective_target=Decimal("45.00"),
        steps=(
            QuoteStep(
                round_no=1,
                start_amount=Decimal("100.00"),
                end_amount=Decimal("55.00"),
                reduction_amount=Decimal("45.00"),
                reduction_pct=Decimal("45.00"),
            ),
            QuoteStep(
                round_no=2,
                start_amount=Decimal("55.00"),
                end_amount=Decimal("45.00"),
                reduction_amount=Decimal("10.00"),
                reduction_pct=Decimal("18.18"),
            ),
        ),
    )


def test_write_new_file(tmp_path: Path) -> None:
    out = tmp_path / "nested" / "dir" / "quotes.xlsx"
    write_result(_result(_params(), datetime(2026, 4, 19, 10, 30, 0)), out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["2026-04-19"]
    ws = wb["2026-04-19"]
    assert ws.cell(1, 1).value == "Overview"
    assert ws.cell(2, 2).value == "2026-04-19 10:30:00"
    assert ws.cell(3, 2).value == "100"
    assert ws.cell(10, 1).value == "Round"
    assert ws.cell(11, 2).value == "100.00"


def test_write_same_day_prepends(tmp_path: Path) -> None:
    out = tmp_path / "quotes.xlsx"
    write_result(_result(_params(), datetime(2026, 4, 19, 10, 30, 0)), out)
    write_result(_result(_params(), datetime(2026, 4, 19, 11, 45, 0)), out)
    wb = load_workbook(out)
    ws = wb["2026-04-19"]
    assert ws.cell(1, 1).value == "Overview"
    assert ws.cell(2, 2).value == "2026-04-19 11:45:00"
    # Block = 10 header rows + 2 step rows = 12. Then 2 blank rows, then prev.
    assert ws.cell(13, 1).value is None
    assert ws.cell(14, 1).value is None
    assert ws.cell(15, 1).value == "Overview"
    assert ws.cell(16, 2).value == "2026-04-19 10:30:00"


def test_write_new_day_new_sheet(tmp_path: Path) -> None:
    out = tmp_path / "quotes.xlsx"
    write_result(_result(_params(), datetime(2026, 4, 19, 10, 0)), out)
    write_result(_result(_params(), datetime(2026, 4, 20, 10, 0)), out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"2026-04-19", "2026-04-20"}


# --- main -------------------------------------------------------------------


def _argv(tmp_path: Path, **overrides) -> list[str]:
    base = {
        "--start-price": "100",
        "--target-price": "45",
        "--max-pct": "50",
        "--min-reduction": "10",
        "--decimals": "2",
        "--rounding": "true",
        "--output": str(tmp_path / "quotes.xlsx"),
    }
    base.update(overrides)
    out: list[str] = []
    for k, v in base.items():
        out += [k, v]
    return out


def test_main_success(tmp_path: Path) -> None:
    assert main(_argv(tmp_path)) == 0
    assert (tmp_path / "quotes.xlsx").exists()


def test_main_missing_arg(tmp_path: Path) -> None:
    argv = _argv(tmp_path)
    i = argv.index("--rounding")
    del argv[i : i + 2]
    with pytest.raises(SystemExit) as exc:
        main(argv)
    assert exc.value.code == 2


def test_main_validation_error(tmp_path: Path) -> None:
    assert main(_argv(tmp_path, **{"--start-price": "0"})) == 3


def test_main_infeasible(tmp_path: Path) -> None:
    assert (
        main(
            _argv(
                tmp_path,
                **{
                    "--start-price": "10",
                    "--target-price": "5",
                    "--max-pct": "10",
                    "--min-reduction": "10",
                },
            )
        )
        == 4
    )
